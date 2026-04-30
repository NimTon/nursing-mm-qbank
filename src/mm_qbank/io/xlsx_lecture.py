from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 与 refine 的流式 CSV 同风格，便于断点续跑（行号 = Excel 数据行号 2,3,4…）
LECTURE_TIPS_STREAM_COLUMNS: tuple[str, ...] = (
    "行号",
    "题号",
    "讲师提醒",
)


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def read_result_xlsx(path: Path) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    """
    读取首个工作表首行为表头；自第 2 行起为数据行。
    返回 (表头, [(excel_row, row_dict), ...])，row_dict 键为表头字符串，值为单元格文本。
    """
    path = path.resolve()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return ([], [])
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return ([], [])
        headers: list[str] = []
        for i, c in enumerate(header_row):
            h = _cell_str(c) or f"列{i + 1}"
            headers.append(h)
        if not headers:
            return ([], [])

        out: list[tuple[int, dict[str, str]]] = []
        for r_idx, row in enumerate(rows_iter, start=2):
            d: dict[str, str] = {}
            for j, h in enumerate(headers):
                v = row[j] if j < len(row) else None
                d[h] = _cell_str(v)
            out.append((r_idx, d))
        return (headers, out)
    finally:
        wb.close()


def write_lecture_tips_xlsx(
    path: Path,
    headers: Sequence[str],
    data_rows: list[dict[str, str]],
) -> None:
    """将含「讲师提醒」等列的 dict 行写入新 xlsx（表头 = headers 顺序）。"""
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    hb = [str(h) for h in headers]
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = "lecture_tips"
    for c, h in enumerate(hb, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, d in enumerate(data_rows, start=2):
        for c, h in enumerate(hb, start=1):
            v = d.get(h, "")
            if v is None:
                v = ""
            ws.cell(row=r, column=c, value=v)
    for c in range(1, len(hb) + 1):
        w = 14 if c == 1 else min(50, 18)
        ws.column_dimensions[get_column_letter(c)].width = w
    wb.save(path)


def append_lecture_tips_rows_to_csv(path: Path, rows: list[dict[str, Any]]) -> int:
    """流式 UTF-8-SIG 追加；列固定为 LECTURE_TIPS_STREAM_COLUMNS。"""
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = LECTURE_TIPS_STREAM_COLUMNS
    needs_header = not path.exists() or path.stat().st_size <= 0
    w = 0
    with path.open("a", encoding="utf-8-sig", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(columns), extrasaction="ignore")
        if needs_header:
            wr.writeheader()
        for row in rows:
            o = {k: ("" if row.get(k) is None else str(row.get(k) or "")) for k in columns}
            wr.writerow(o)
            w += 1
        f.flush()
    return w


def load_done_行号_from_lecture_csv(path: Path) -> set[int]:
    """从已有流式 CSV 读取已完成的行号，用于断点续跑。"""
    p = path.resolve()
    if not p.exists():
        return set()
    done: set[int] = set()
    try:
        with p.open("r", encoding="utf-8-sig", newline="") as f:
            r = csv.DictReader(f)
            for row in r:
                if not row:
                    continue
                s = str(row.get("行号", "") or "").strip()
                if s.isdigit():
                    done.add(int(s))
    except Exception:  # noqa: BLE001
        return set()
    return done
