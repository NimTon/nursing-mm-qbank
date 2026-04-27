from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 表头与行 dict 键一致
_BASE_REFINED_XLSX_COLUMNS: Sequence[str] = (
    "题号",
    "修正状态",
    "原问题",
    "原解析",
    "修正后问题",
    "修正问题原因",
    "修正问题参考来源",
    "修正后解析",
    "修正解析原因",
    "修正解析参考来源",
    "讲师提醒",
)


def _as_修正状态_cell(v: Any) -> str:
    if v is True:
        return "是"
    if v is False:
        return "否"
    if isinstance(v, str) and v.strip() in ("是", "已修正", "true", "True", "1"):
        return "是"
    if isinstance(v, str) and v.strip() in ("否", "未修正", "false", "0"):
        return "否"
    return "否"


def write_refined_rows_to_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = "refined"
    columns = tuple(_BASE_REFINED_XLSX_COLUMNS)
    for c, h in enumerate(columns, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(columns, start=1):
            v = row.get(key, "")
            if v is None:
                v = ""
            if key == "修正状态":
                ws.cell(row=r, column=c, value=_as_修正状态_cell(v))
            else:
                ws.cell(row=r, column=c, value=v)
    for c in range(1, len(columns) + 1):
        w = 14 if c <= 2 else (min(50, 22) if c == len(columns) else min(50, 14))
        ws.column_dimensions[get_column_letter(c)].width = w
    wb.save(path)
