from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from mm_qbank.io.xlsx_out import write_refined_rows_to_xlsx


def _sample_row() -> dict[str, object]:
    return {
        "page_id": "p1",
        "source_image": "x.png",
        "题号": "1",
        "题目类型": "",
        "原问题": "Q",
        "原解析": "A0",
        "修正后问题": "Q2",
        "修正问题原因": "r1",
        "修正问题参考来源": "b1",
        "修正后解析": "A1",
        "修正解析原因": "r2",
        "修正解析参考来源": "b2",
        "讲师提醒": "",
        "要点": "",
        "讲课内容": "",
        "修正状态": True,
    }


def test_write_refined_xlsx_header(tmp_path: Path) -> None:
    p = tmp_path / "a.xlsx"
    write_refined_rows_to_xlsx(p, [_sample_row()])
    assert p.is_file()
    wb = load_workbook(p)
    ws = wb.active
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "讲师提醒" in hdr
    assert "要点" in hdr
    assert "讲课内容" in hdr


def test_write_refined_xlsx_omit_tips_and_lecture_columns(tmp_path: Path) -> None:
    p = tmp_path / "b.xlsx"
    write_refined_rows_to_xlsx(
        p,
        [_sample_row()],
        include_lecture_tips=False,
        include_lecture_content=False,
    )
    wb = load_workbook(p)
    ws = wb.active
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "讲师提醒" not in hdr
    assert "要点" not in hdr
    assert "讲课内容" not in hdr
