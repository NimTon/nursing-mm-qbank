from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from mm_qbank.io.xlsx_lecture import (
    append_lecture_tips_rows_to_csv,
    load_done_行号_from_lecture_csv,
    read_result_xlsx,
    write_lecture_tips_xlsx,
)
from mm_qbank.pipeline.lecture_tips_run import _build_llm_items, _merge_tips_to_rows


def test_read_write_lecture_xlsx(tmp_path: Path) -> None:
    p = tmp_path / "t.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise AssertionError("no active sheet")
    ws.cell(1, 1, "题号")
    ws.cell(1, 2, "原问题")
    ws.cell(1, 3, "原解析")
    ws.cell(2, 1, "1")
    ws.cell(2, 2, "Q")
    ws.cell(2, 3, "A")
    wb.save(p)

    headers, rows = read_result_xlsx(p)
    assert "题号" in headers
    assert len(rows) == 1
    r, d = rows[0]
    assert r == 2
    assert d["题号"] == "1"
    h2, merged = _merge_tips_to_rows(
        headers,
        rows,
        {2: "（提示）"},
    )
    assert "讲师提醒" in h2
    assert merged[0]["讲师提醒"] == "（提示）"
    out = tmp_path / "o.xlsx"
    write_lecture_tips_xlsx(out, h2, merged)
    h3, again = read_result_xlsx(out)
    assert h3[0] == h2[0]
    assert again[0][1].get("讲师提醒") == "（提示）"


def test_lecture_csv_append_and_done(tmp_path: Path) -> None:
    c = tmp_path / "s.csv"
    n = append_lecture_tips_rows_to_csv(c, [{"行号": 2, "题号": "1", "讲师提醒": "x"}])
    assert n == 1
    done = load_done_行号_from_lecture_csv(c)
    assert 2 in done


def test_build_llm_items_row_numbers() -> None:
    b = _build_llm_items(
        [
            (
                2,
                {"题号": "1", "原问题": "P", "原解析": "R", "其他": ""},
            )
        ]
    )
    assert b[0]["行号"] == 2
    assert b[0]["题号"] == "1"
    assert "问题" in b[0]
